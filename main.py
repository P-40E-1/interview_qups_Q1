from selenium import webdriver
from selenium.webdriver.common.by import By
from datetime import datetime
import openpyxl
import time
import re

# Define the path to the Chrome WebDriver executable
webdriver_path = r"C:\SeleniumDrivers\chromedriver.exe"

# Initialize the WebDriver
options = webdriver.ChromeOptions()
options.add_argument("--lang=en")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome()

current_day = datetime.now().strftime("%A")

try:
    # Load the Excel workbook
    workbook = openpyxl.load_workbook("Excel.xlsx")
    sheet_names = workbook.sheetnames
    matching_sheets = [sheet_name for sheet_name in sheet_names if sheet_name == current_day]

    if matching_sheets:
        sheet = workbook[matching_sheets[0]]  # Select the matching sheet

        for row in range(2, sheet.max_row + 1):  # Start from row 2 (assuming row 1 is the header)
            search_term = sheet.cell(row=row, column=3).value

            if search_term:
                driver.get("https://www.google.com/en")
                search_input = driver.find_element(By.NAME, 'q')
                search_input.clear()
                search_input.send_keys(search_term)
                time.sleep(2)

                # Extract and put English suggestions in columns 4 and 5
                suggestions = driver.find_elements(By.XPATH, "//ul[@role='listbox']//li")
                english_suggestions = []

                for suggestion in suggestions:
                    text = suggestion.text
                    # Check if the suggestion contains English characters using a regular expression
                    if re.match(r'^[a-zA-Z\s]+$', text):
                        english_suggestions.append(text)

                if english_suggestions:
                    # Sort English suggestions by length (shortest to longest)
                    english_suggestions.sort(key=lambda x: len(x))

                    sheet.cell(row=row, column=5).value = english_suggestions[0]  # Shortest English suggestion
                    sheet.cell(row=row, column=4).value = english_suggestions[-1]  # Longest English suggestion

    # Save the updated Excel file
    workbook.save("Excel.xlsx")

except Exception as e:
    print(f"An error occurred: {str(e)}")

finally:
    driver.quit()  # Close the WebDriver when finished
